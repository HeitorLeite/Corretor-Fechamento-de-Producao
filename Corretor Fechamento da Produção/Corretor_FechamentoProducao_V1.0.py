"""
Conversor de dados do Excel para formato CSV
Lê a planilha de Eventos para Fechamento da Produção dos Prestadores
e transforma em arquivo CSV no formato esperado

Desenvolvido por: Heitor Leite - 2026
"""

import openpyxl
import os
import sys
from pathlib import Path

def processar_planilha_eventos(caminho_xlsx=None, arquivo_saida=None):
    """
    Processa a planilha de eventos e gera arquivo CSV com os dados estruturados
    Estrutura do CSV:
    - Coluna 1: Código do Cooperado/Prestador
    - Coluna 2: Código do Cooperado/Prestador (duplicado)
    - Coluna 3: Valor Lançamento PJ (em branco se não existir)
    - Coluna 4: Valor Lançamento PF (em branco se não existir)
    - Coluna 5: Código da Cidade de Atendimento

    Os parâmetros opcionais permitem indicar um arquivo de origem e o arquivo
    de destino. Se não forem fornecidos, o usuário será solicitado a digitá-los.
    Para remover a quinta coluna do CSV (cidade), veja comentário na montagem
    da linha.
    """
    
    # Se nenhum caminho for passado, pergunta ao usuário usando input
    if caminho_xlsx is None:
        caminho_xlsx = input('Entre com o caminho do arquivo Excel (ou arraste e solte aqui): ').strip()

    # limpar eventuais aspas, & e espaços extras que o PowerShell/pasta copiar podem adicionar
    if caminho_xlsx.startswith('&'):
        # & pode aparecer quando se cola do PowerShell
        caminho_xlsx = caminho_xlsx.lstrip('&').strip()
    caminho_xlsx = caminho_xlsx.strip('"').strip("'")

    # Caminho do arquivo Excel informado agora em caminho_xlsx
    # Validar se o arquivo existe
    if not caminho_xlsx or not os.path.exists(caminho_xlsx):
        print(f'ERRO: Arquivo não encontrado em {caminho_xlsx}')
        return False
    
    print(f'Lendo arquivo: {caminho_xlsx}')
    
    try:
        # Carregar o arquivo Excel
        wb = openpyxl.load_workbook(caminho_xlsx)
        ws = wb.active
        
        # Lista para armazenar os dados processados
        dados_processados = []
        
        # Encontrar todas as seções de dados
        # Uma seção começa após "Cod Cooperado/Prestador" e vai até uma linha vazia ou outra seção
        secoes_dados = []
        for linha_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row[0] and isinstance(row[0], str):
                conteudo = str(row[0]).lower()
                if 'cod' in conteudo and 'cooperado' in conteudo:
                    # Encontrou um cabeçalho, a seção começa na próxima linha
                    linha_inicio = linha_idx + 1
                    
                    # Encontrar onde essa seção termina (próxima linha vazia ou fim do arquivo)
                    linha_fim = ws.max_row + 1
                    for linha_check in range(linha_inicio, ws.max_row + 1):
                        row_check = list(ws.iter_rows(min_row=linha_check, max_row=linha_check, values_only=True))[0]
                        # Se encontrar uma linha vazia (ambas colunas vazias)
                        if row_check[0] is None and (len(row_check) < 2 or row_check[1] is None):
                            linha_fim = linha_check
                            break
                        # Se encontrar outro cabeçalho "Cod Cooperado/Prestador"
                        if row_check[0] and isinstance(row_check[0], str):
                            conteudo_check = str(row_check[0]).lower()
                            if 'cod' in conteudo_check and 'cooperado' in conteudo_check:
                                linha_fim = linha_check
                                break
                    
                    secoes_dados.append((linha_inicio, linha_fim))
                    print(f'Seção de dados encontrada: linhas {linha_inicio} a {linha_fim - 1}')
        
        if not secoes_dados:
            # Se não encontrar nenhuma seção, assume linha 6 (padrão)
            secoes_dados = [(6, ws.max_row + 1)]
            print(f'Nenhuma seção encontrada. Usando linha 6 como padrão.')
        
        print(f'   Total de seções encontradas: {len(secoes_dados)}')
        
        # Palavras-chave que indicam linhas de cabeçalho/separador
        # Apenas termos que aparecem realmente em cabeçalhos, não em nomes de empresas
        palavras_chave_cabecalho = [
            'desconto producao',  # DESCONTO PRODUCAO/GLOSA
            'nome evento',
            'nome prestd',  # Pode aparecer como "Nome Prestdor" com variações
        ]
        
        # Processar cada linha de dados
        total_linhas = 0
        # Processar todas as seções encontradas
        for idx_secao, (linha_inicio, linha_fim) in enumerate(secoes_dados):
            print(f'\n[SECAO {idx_secao + 1}] Processando linhas {linha_inicio} a {linha_fim - 1}...')
            
            for row in ws.iter_rows(min_row=linha_inicio, max_row=linha_fim - 1, values_only=True):
                # Coluna 1: Código do Cooperado/Prestador
                cod_cooperado = row[0]
                
                # Pular linhas vazias ou sem código
                if cod_cooperado is None:
                    continue
                
                # Converter para string e fazer strip
                cod_str = str(cod_cooperado).strip()
                
                # Se for vazio após strip, pular
                if not cod_str:
                    continue
                
                # Verificar se é número
                try:
                    cod_cooperado_int = int(float(cod_str))
                except (ValueError, TypeError):
                    # Se não for número, pular este item
                    continue
                
                # Verificar se a segunda coluna é um cabeçalho
                segunda_coluna = row[1] if len(row) > 1 else None
                if segunda_coluna is not None:
                    segunda_str = str(segunda_coluna).strip().lower()
                    # Se contiver palavras-chave de cabeçalho, pular a linha
                    if any(palavra in segunda_str for palavra in palavras_chave_cabecalho):
                        continue
                
                # Coluna 3: Valor Lançamento PJ
                valor_lancamento_pj = row[2] if len(row) > 2 else None
                valor_pj_str = ''
                if valor_lancamento_pj is not None:
                    valor_pj_str = str(valor_lancamento_pj).strip() if valor_lancamento_pj else ''
                
                # Coluna 4: Valor Lançamento PF
                valor_lancamento_pf = row[3] if len(row) > 3 else None
                valor_pf_str = ''
                if valor_lancamento_pf is not None:
                    valor_pf_str = str(valor_lancamento_pf).strip() if valor_lancamento_pf else ''
                
                # Coluna 5: Código da Cidade de Atendimento
                cidade_atendimento = row[4] if len(row) > 4 else None
                cidade_str = ''
                if cidade_atendimento is not None:
                    cidade_str = str(cidade_atendimento).strip() if cidade_atendimento else ''
                
                # Montar a linha no formato:
                # codigo;codigo_duplicado;valor_pj;valor_pf;cidade
                # Para retirar a quinta coluna e gerar apenas 4 campos, basta
                # comentar a parte ";{cidade_str}" abaixo ou montar a string
                # manualmente com apenas os quatro primeiros valores.
                # Exemplo para 4 colunas:
                linha = f"{cod_cooperado_int};{cod_cooperado_int};{valor_pj_str};{valor_pf_str}"
                #linha = f'{cod_cooperado_int};{cod_cooperado_int};{valor_pj_str};{valor_pf_str};{cidade_str}'
                dados_processados.append(linha)
                
                total_linhas += 1
                print(f'  Linha {total_linhas}: {linha}')
        
        # Validar se encontrou dados
        if not dados_processados:
            print('Nenhum dado foi processado. Verifique a planilha.')
            return False
        
        # Se não foi passado um caminho de saída, perguntar para o usuário
        if arquivo_saida is None:
            # sugestão: informe caminho completo ou apenas nome; se quiser pasta diferente, inclua
            arquivo_saida = input('Entre com o caminho/arquivo de saída CSV (ex: Saidas/saida.csv): ').strip()
            if arquivo_saida.startswith('&'):
                arquivo_saida = arquivo_saida.lstrip('&').strip()
            arquivo_saida = arquivo_saida.strip('"').strip("'")
            if not arquivo_saida:
                # uso padrão em Saidas/
                diretorio_saida = 'Saidas'
                Path(diretorio_saida).mkdir(exist_ok=True)
                arquivo_saida = os.path.join(diretorio_saida, 'fechamento_producao_saida.csv')
            else:
                # garantir diretório existe
                pasta = os.path.dirname(arquivo_saida)
                if pasta:
                    Path(pasta).mkdir(parents=True, exist_ok=True)

        # Se o usuário informou apenas uma pasta (ou a pasta existe)
        if os.path.isdir(arquivo_saida) or arquivo_saida.endswith(os.sep):
            diretorio_saida = arquivo_saida.rstrip(os.sep)
            Path(diretorio_saida).mkdir(parents=True, exist_ok=True)
            arquivo_saida = os.path.join(diretorio_saida, 'fechamento_producao_saida.csv')
        else:
            # criar diretório de saída se não existir (caso tenha sido passado apenas nome)
            diretorio_saida = os.path.dirname(arquivo_saida) or ''
            if diretorio_saida:
                Path(diretorio_saida).mkdir(exist_ok=True)
        
        print(f'\nSALVANDO {len(dados_processados)} linhas em {arquivo_saida}...')
        
        try:
            with open(arquivo_saida, 'w', encoding='utf-8', newline='') as f:
                for linha in dados_processados:
                    f.write(linha + '\n')
            
            print(f'Arquivo salvo com sucesso!')
            print(f'\n=== INFORMAÇÕES ===')
            print(f'Total de linhas processadas: {len(dados_processados)}')
            print(f'Arquivo de saída: {arquivo_saida}')
            return True
            
        except Exception as e:
            print(f'Erro ao salvar arquivo: {e}')
            return False
    
    except Exception as e:
        print(f'ERRO ao processar planilha: {e}')
        return False


def main():
    """Função principal"""
    print('=' * 60)
    print('CONVERSOR DE PLANILHA - FECHAMENTO DA PRODUÇÃO')
    print('=' * 60 + '\n')

    # permitir que o usuário passe caminho de origem e destino como argumentos
    # Exemplo de uso: python script.py "Planilhas/meuarquivo.xlsx" "Saidas/saida.csv"
    caminho = None
    saida = None
    if len(sys.argv) > 1:
        caminho = sys.argv[1]
    if len(sys.argv) > 2:
        saida = sys.argv[2]

    sucesso = processar_planilha_eventos(caminho, saida)
    
    print('\n' + '=' * 60)
    if sucesso:
        print('Processo concluído com sucesso!')
    else:
        print('Processo concluído com erros!')
    print('=' * 60)


if __name__ == '__main__':
    main()