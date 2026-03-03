"""
Conversor de dados do Excel para formato CSV com Interface Gráfica

Lê a planilha de Eventos para Fechamento da Produção dos Prestadores
converte para CSV no formato esperado.

Desenvolvido por: Heitor Leite - 2026
Interface gráfica integrada em 2026
"""

# =============================
# IMPORTAÇÕES
# =============================

# Biblioteca para trabalhar com Excel
import openpyxl

# Biblioteca para arquivos e diretórios
import os

# Biblioteca para argumentos no terminal
import sys

# Biblioteca moderna para caminhos
from pathlib import Path

# Bibliotecas para Interface Gráfica
import tkinter as tk
from tkinter import ttk, filedialog, messagebox


# =============================
# FUNÇÃO DE PROCESSAMENTO
# =============================

def processar_planilha_eventos(caminho_xlsx, arquivo_saida, callback_log=None):
    """
    Processa a planilha Excel e gera o CSV.

    Parâmetros:
    - caminho_xlsx: caminho do arquivo Excel
    - arquivo_saida: caminho do CSV
    - callback_log: função para mostrar logs na interface
    """

    # Função interna para registrar mensagens
    def log(msg):
        if callback_log:
            callback_log(msg)
        else:
            print(msg)

    # =============================
    # VALIDAÇÃO DO ARQUIVO
    # =============================

    if not caminho_xlsx or not os.path.exists(caminho_xlsx):
        log(f"ERRO: Arquivo não encontrado: {caminho_xlsx}")
        return False

    log(f"Lendo arquivo: {caminho_xlsx}")

    try:

        # =============================
        # ABERTURA DO EXCEL
        # =============================

        wb = openpyxl.load_workbook(caminho_xlsx)
        ws = wb.active

        # Lista que armazenará os dados finais
        dados_processados = []

        # =============================
        # LOCALIZAÇÃO DAS SEÇÕES
        # =============================

        secoes_dados = []

        for linha_idx, row in enumerate(ws.iter_rows(values_only=True), 1):

            if row[0] and isinstance(row[0], str):

                conteudo = str(row[0]).lower()

                if 'cod' in conteudo and 'cooperado' in conteudo:

                    linha_inicio = linha_idx + 1
                    linha_fim = ws.max_row + 1

                    for linha_check in range(linha_inicio, ws.max_row + 1):

                        row_check = list(
                            ws.iter_rows(
                                min_row=linha_check,
                                max_row=linha_check,
                                values_only=True
                            )
                        )[0]

                        # Linha vazia encerra seção
                        if row_check[0] is None and (
                            len(row_check) < 2 or row_check[1] is None
                        ):
                            linha_fim = linha_check
                            break

                        # Novo cabeçalho encerra seção
                        if row_check[0] and isinstance(row_check[0], str):

                            conteudo_check = str(row_check[0]).lower()

                            if 'cod' in conteudo_check and 'cooperado' in conteudo_check:
                                linha_fim = linha_check
                                break

                    secoes_dados.append((linha_inicio, linha_fim))

                    log(f"Seção encontrada: {linha_inicio} até {linha_fim - 1}")

        # Caso não encontre seções
        if not secoes_dados:

            secoes_dados = [(6, ws.max_row + 1)]
            log("Nenhuma seção encontrada. Usando linha 6.")

        log(f"Total de seções: {len(secoes_dados)}")

        # =============================
        # FILTROS DE CABEÇALHO
        # =============================

        palavras_chave_cabecalho = [
            'desconto producao',
            'nome evento',
            'nome prestd'
        ]

        total_linhas = 0

        # =============================
        # PROCESSAMENTO DOS DADOS
        # =============================

        for idx, (inicio, fim) in enumerate(secoes_dados, 1):

            log(f"Processando seção {idx}: {inicio} até {fim - 1}")

            for row in ws.iter_rows(
                min_row=inicio,
                max_row=fim - 1,
                values_only=True
            ):

                # Código do cooperado
                cod = row[0]

                if cod is None:
                    continue

                cod_str = str(cod).strip()

                if not cod_str:
                    continue

                try:
                    cod_int = int(float(cod_str))
                except:
                    continue

                # Verifica cabeçalho na segunda coluna
                segunda = row[1] if len(row) > 1 else None

                if segunda:

                    segunda_str = str(segunda).lower().strip()

                    if any(p in segunda_str for p in palavras_chave_cabecalho):
                        continue

                # Valores
                valor_pj = row[2] if len(row) > 2 else None
                valor_pf = row[3] if len(row) > 3 else None
                cidade = row[4] if len(row) > 4 else None

                valor_pj_str = str(valor_pj).strip() if valor_pj else ''
                valor_pf_str = str(valor_pf).strip() if valor_pf else ''
                cidade_str = str(cidade).strip() if cidade else ''

                # Montagem da linha CSV
                linha = (
                    f"{cod_int};{cod_int};"
                    f"{valor_pj_str};"
                    f"{valor_pf_str};"
                    f"{cidade_str}"
                )

                dados_processados.append(linha)
                total_linhas += 1

        # =============================
        # VALIDAÇÃO
        # =============================

        if not dados_processados:
            log("Nenhum dado encontrado.")
            return False

        # =============================
        # CRIAÇÃO DE PASTAS
        # =============================

        pasta = os.path.dirname(arquivo_saida)

        if pasta:
            Path(pasta).mkdir(parents=True, exist_ok=True)

        # =============================
        # GRAVAÇÃO DO CSV
        # =============================

        log(f"Salvando arquivo: {arquivo_saida}")

        with open(arquivo_saida, 'w', encoding='utf-8', newline='') as f:

            for linha in dados_processados:
                f.write(linha + '\n')

        log("Arquivo salvo com sucesso!")
        log(f"Total de linhas: {len(dados_processados)}")

        return True

    except Exception as e:

        log(f"Erro no processamento: {e}")
        return False


# =============================
# CLASSE DA INTERFACE GRÁFICA
# =============================

class Aplicacao(tk.Tk):

    def __init__(self):

        super().__init__()

        # Configuração da janela
        self.title("Conversor Excel → CSV")
        self.geometry("720x550")
        self.resizable(False, False)
        self.configure(bg="#f2f4f8")

        # Criação do estilo
        self.criar_estilo()

        # Criação dos componentes
        self.criar_widgets()


    # =============================
    # ESTILO VISUAL
    # =============================

    def criar_estilo(self):

        style = ttk.Style()
        style.theme_use("clam")

        style.configure(
            "TButton",
            font=("Segoe UI", 11),
            padding=8
        )

        style.configure(
            "TEntry",
            font=("Segoe UI", 11),
            padding=6
        )

        style.configure(
            "TLabel",
            font=("Segoe UI", 11),
            background="#f2f4f8"
        )

        style.configure(
            "Titulo.TLabel",
            font=("Segoe UI", 20, "bold"),
            background="#f2f4f8"
        )


    # =============================
    # COMPONENTES DA TELA
    # =============================

    def criar_widgets(self):

        container = tk.Frame(self, bg="#f2f4f8")
        container.pack(fill="both", expand=True, padx=30, pady=20)

        # Título
        ttk.Label(
            container,
            text="Conversor de Planilhas",
            style="Titulo.TLabel"
        ).pack(pady=10)

        ttk.Label(
            container,
            text="Converta arquivos Excel em CSV"
        ).pack(pady=5)

        # =============================
        # ENTRADA
        # =============================

        frame_in = tk.LabelFrame(
            container,
            text=" Arquivo de Entrada ",
            bg="#f2f4f8",
            font=("Segoe UI", 11, "bold")
        )
        frame_in.pack(fill="x", pady=15)

        self.var_entrada = tk.StringVar()

        ttk.Entry(
            frame_in,
            textvariable=self.var_entrada,
            width=60
        ).pack(side="left", padx=10, pady=10)

        ttk.Button(
            frame_in,
            text="Procurar",
            command=self.buscar_entrada
        ).pack(side="right", padx=10)

        # =============================
        # SAÍDA
        # =============================

        frame_out = tk.LabelFrame(
            container,
            text=" Arquivo de Saída ",
            bg="#f2f4f8",
            font=("Segoe UI", 11, "bold")
        )
        frame_out.pack(fill="x", pady=10)

        self.var_saida = tk.StringVar()

        ttk.Entry(
            frame_out,
            textvariable=self.var_saida,
            width=60
        ).pack(side="left", padx=10, pady=10)

        ttk.Button(
            frame_out,
            text="Salvar Como",
            command=self.buscar_saida
        ).pack(side="right", padx=10)

        # =============================
        # BOTÃO CONVERTER
        # =============================

        self.btn_converter = ttk.Button(
            container,
            text="▶ Converter",
            command=self.executar_conversao
        )

        self.btn_converter.pack(pady=20)

        # =============================
        # LOG
        # =============================

        frame_log = tk.LabelFrame(
            container,
            text=" Log ",
            bg="#f2f4f8",
            font=("Segoe UI", 11, "bold")
        )
        frame_log.pack(fill="both", expand=True)

        self.txt_log = tk.Text(
            frame_log,
            height=10,
            font=("Consolas", 10),
            bg="#1e1e1e",
            fg="#00ff9c"
        )

        self.txt_log.pack(fill="both", expand=True, padx=10, pady=10)

        self.log("Sistema pronto para uso.\n")


    # =============================
    # FUNÇÕES DA INTERFACE
    # =============================

    def log(self, msg):

        self.txt_log.insert(tk.END, msg + "\n")
        self.txt_log.see(tk.END)
        self.update()


    def buscar_entrada(self):

        arquivo = filedialog.askopenfilename(
            title="Selecionar Excel",
            filetypes=[("Excel", "*.xlsx")]
        )

        if arquivo:

            self.var_entrada.set(arquivo)

            pasta = os.path.dirname(arquivo)
            nome = os.path.splitext(os.path.basename(arquivo))[0]

            sugestao = os.path.join(
                pasta,
                f"{nome}_convertido.csv"
            )

            self.var_saida.set(sugestao)


    def buscar_saida(self):

        arquivo = filedialog.asksaveasfilename(
            title="Salvar CSV",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")]
        )

        if arquivo:
            self.var_saida.set(arquivo)


    def executar_conversao(self):

        entrada = self.var_entrada.get()
        saida = self.var_saida.get()

        if not entrada:
            messagebox.showwarning("Atenção", "Selecione o Excel.")
            return

        if not saida:
            messagebox.showwarning("Atenção", "Selecione o CSV.")
            return

        self.txt_log.delete(1.0, tk.END)
        self.log("Iniciando conversão...")

        self.btn_converter.config(state="disabled")

        sucesso = processar_planilha_eventos(
            entrada,
            saida,
            callback_log=self.log
        )

        self.btn_converter.config(state="normal")

        if sucesso:

            messagebox.showinfo("Sucesso", "Conversão concluída!")
            self.log("Processo finalizado com êxito.")

        else:

            messagebox.showerror("Erro", "Falha na conversão.")
            self.log("Processo finalizado com erro.")


# =============================
# EXECUÇÃO
# =============================

if __name__ == '__main__':

    app = Aplicacao()
    app.mainloop()
